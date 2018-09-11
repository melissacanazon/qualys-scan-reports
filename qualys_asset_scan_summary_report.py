##############################################################
# Generate Qualys Scan Summary Report                        #
# Date: 06/19/2018                                           #
# Version 1.0                                                #
##############################################################

###############
### MODULES ###
###############
import argparse
import collections
from datetime import datetime
from datetime import timedelta
import ipaddress
import json
from lxml import etree as ET
import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import pytz
import re
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from smb.SMBConnection import SMBConnection
import timeit
import zipfile

###############################
### ARGUMENTS AND VARIABLES ###
###############################

## PARSE COMMANDLINE ARGUMENTS ##
parser = argparse.ArgumentParser()
parser.add_argument('-D', '--Debug', help='Debug Mode assists in determining issues being raised by the script.', action='store_true')
parser.add_argument('-RD', '--ReportDate', type=lambda s: datetime.strptime(s, '%Y%m%d').replace(hour=23, minute=59, second=59, microsecond=999999), default=datetime.today().replace(hour=23, minute=59, second=59, microsecond=999999), help='Specify the report date in YYYYMMDD format.')
parser.add_argument('-LB', '--LaunchedBefore', type=int, default=30, help='Days - Scan Launched Before.')
parser.add_argument('-LA', '--LaunchedAfter', type=int, default=30, help='Days - Scan Launched After.')
parser.add_argument('-IPL', '--IPList', help='IP listing of scan targets.', default='')
parser.add_argument('-FNP', '--FileNamePrefix', help='Report Filename Prefix.')
args = parser.parse_args()


## INITIALIZE REPORT SERVER CONNECTION VARIABLES ##
ReportServer = 'yourServer'
ReportServerIP = '192.168.1.2'
ReportUsername = 'report_user@reporttothe.ninja'
ReportPassword = 'password or open(path/password).read()' 

## INITIALIZE VULNERABILITY SCANNER (Qualys) CONNECTION VARIABLES ##
VSServer = 'https://qualysapi.qualys.com' 
VSUrl = VSServer + '/api/2.0/fo/' 
VSUsername = 'vs_ninja' 
VSPassword = 'password or open(path/password).read()'
VSVerify_SSL = False
VSSession = requests.Session()
VSSession.headers.update({'X-Requested-With': 'VS User'})

report_file_name = args.FileNamePrefix + '_Qualys_Scan_Summary_Report_' + args.ReportDate.strftime("%Y%m%d") + '.xlsx'
report_file_path = os.path.join('report', 'file_path', report_file_name)
report_network_path = os.path.join('file', 'path', 'on_share', report_file_name)

report_fields = collections.OrderedDict()
report_fields.update({'ip' : {'field_category' : 'vuln', 'column_header' : 'IP'}})
report_fields.update({'scanTitle' : {'field_category' : 'vuln', 'column_header' : 'Scan Title'}})
report_fields.update({'startTimeGMT' : {'field_category' : 'vuln', 'column_header' : 'Start Time (GMT)'}})
report_fields.update({'endTimeGMT' : {'field_category' : 'vuln', 'column_header' : 'End Time (GMT)'}})
report_fields.update({'startTimeEDT' : {'field_category' : 'vuln', 'column_header' : 'Start Time (EDT)'}})
report_fields.update({'endTimeEDT' : {'field_category' : 'vuln', 'column_header' : 'End Time (EDT)'}})
report_fields.update({'duration' : {'field_category' : 'vuln', 'column_header' : 'Duration'}})
#report_fields.update({'scanReferenceID' : {'field_category' : 'vuln', 'column_header' : 'Scan Reference ID'}})

lAssetIPs = args.IPList
lInfoVulnQIDs = ['45038']

########################
### HELPER FUNCTIONS ###
########################

def VSConnect(sMethod='get', resource='', action='', input={}):
    # SET UP VARIABLES
    dHTTPStatus = {}
    input.update({'action' : action})

    # SEND REQUEST TO QUALYS AND CAPTURE RESPONSE
    try:
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
        oResponse = getattr(VSSession, sMethod)(VSUrl + resource, data=input, verify=VSVerify_SSL)

    except requests.ConnectionError as oErr:
        print('Connection to resource {} - error code {}'.format(resource, str(oErr)))

    return oResponse


def VSLogin():
    # INITIALIZE CONNECTION VARIABLES
    input = {'username': VSUsername,
             'password': VSPassword}

    if (args.Debug):
        print('[+] Debug - Authenticating to Vulnerability Scanner')

    # SUBMIT REQUEST AND PARSE RESPONSE
    VSConnect('post', 'session/', 'login', input)
    
    if (True):
        if (args.Debug):
            print('[+] Debug - Authentication Successful')
    else:
        if (args.Debug):
            print('[+] Debug - Authentication Failed')

# retrieves the all 'SCHEDULED' scans, limited to the scope of the vs user account
def getScans():
    if (args.Debug):
        print('[+] Debug - Retrieving Scan Reference IDs')

    input = {'type' : 'Scheduled',
             'launched_before_datetime' : (datetime.today() - timedelta(days=args.LaunchedBefore)).strftime('%Y-%m-%d'), 
             'launched_after_datetime' : (datetime.today() - timedelta(days=args.LaunchedAfter)).strftime('%Y-%m-%d')} 

    oResponse = VSConnect('post', 'scan/', 'list', input)

    scanReferences = ET.fromstring(oResponse.text.encode('utf-8'))

    lScanReferenceIDs = scanReferences.xpath('.//SCAN')

    if (args.Debug):
        print('Scan Reference IDs: ' + str(len(lScanReferenceIDs)))
    return lScanReferenceIDs


def VSLogout():
    oResponse = VSConnect('post', 'session/', 'logout')
    if (args.Debug):
        print('[+] Debug - Logout Response: ' + oResponse.text)

#####################
### MAIN FUNCTION ###
#####################

# SET TIMER
executionTimerStart = timeit.default_timer()

## CREATE EXCEL REPORT
wb = Workbook()
ws = wb.active

## Open Vuln Details
ws.title = 'Asset Scans'

# WRITE COLUMN HEADERS
for colIndex, field_name in enumerate(report_fields):
    # print(str(colIndex) + ' ' + str(field_name) + ' ' + str(list(report_fields).index('firstSeen')))
    ws.cell(column = colIndex + 1, row = 1, value = report_fields[field_name]['column_header'])

VSLogin()
lScanReferenceIDs = getScans() 

if (args.Debug):
    print('[+] Debug - Retrieving Scan Details')

rowIndex = 1
scanIndex = 0
for scan in lScanReferenceIDs:
    scanIndex += 1
    print('Retrieving Scan Data on ' + str(scanIndex) + ' out of ' + str(len(lScanReferenceIDs)))    

    for element in scan:
        #scan reference id
        if (element.tag == 'REF'):
            scanID = element.text
            print(scanID)
        elif (element.tag =='TITLE'):
            scanTitle = element.text
        elif (element.tag == 'STATUS'):
            for status in element:
                print(status.text)
                # only request and report the scan information if the scan is finished
                if (status.text == 'Finished'):

                    input = {'scan_ref' : scanID,
                             'ips' : ','.join(lAssetIPs), #this could be turned into an arg
                             'output_format' : 'json'}
                                       
                    oResponse = VSConnect('post', 'scan/', 'fetch', input)
                    try: 
                        oResponse.json()
                    except: 
                        print(str(oResponse.text))
                        oResponse = ''
                    else:
                        oResponse = oResponse.json()

                    if oResponse != '':
                        for qid in oResponse:
                            if (qid['qid'] == 45038):
                                rowIndex += 1
                                # the data to report is in one loooong string, lots of splitting
                                result = qid['result'].split('\n\n')
                                for field in result:
                                    gmt = pytz.timezone('GMT')
                                    eastern = pytz.timezone('US/Eastern')
                    
                                    if ('duration' in field):
                                        duration = str(timedelta(seconds=int(field.split(': ')[1].split(' s')[0])))
                                        print(duration)
                                    elif ('Start time' in field):
                                        startGMT = gmt.localize(datetime.strptime(field.split(': ')[1], '%a, %b %d %Y, %H:%M:%S GMT'))
                                        startEDT = startGMT.astimezone(eastern)
                                    elif ('End time' in field):
                                        endGMT = gmt.localize(datetime.strptime(field.split(': ')[1], '%a, %b %d %Y, %H:%M:%S GMT'))
                                        endEDT = endGMT.astimezone(eastern)
                                    else:
                                        print('Unexpected Field Data')
                                for colIndex, field_name in enumerate(report_fields):
                                    if (report_fields[field_name]['field_category'] == 'vuln'):
                                        if (field_name == 'ip'):
                                            cellValue = qid['ip']
                                        elif (field_name == 'scanTitle'):
                                           cellValue = scanTitle
                                        elif (field_name == 'startTimeGMT'):
                                           cellValue = startGMT.strftime('%Y-%m-%d %H:%M:%S')
                                        elif (field_name == 'endTimeGMT'):
                                           cellValue = endGMT.strftime('%Y-%m-%d %H:%M:%S')
                                        elif (field_name == 'startTimeEDT'):
                                           cellValue = startEDT.strftime('%Y-%m-%d %H:%M:%S')
                                        elif (field_name == 'endTimeEDT'):
                                           cellValue = endEDT.strftime('%Y-%m-%d %H:%M:%S') 
                                        elif (field_name == 'duration'):
                                           cellValue = duration
                                    ws.cell(column = colIndex + 1, row = rowIndex, value = cellValue)
                                    if (args.Debug):
                                        print('data written')
                    
wb.save(filename = report_file_path)

## COPY REPORT FILE TO NETWORK SHARE
if (args.Debug):
    print('[+] Debug - Copying report file to network share')

copyFileTimerStart = timeit.default_timer()

conn = SMBConnection(ReportUsername, ReportPassword, 'client_machine_name', ReportServer, use_ntlm_v2 = True)
conn.connect(ReportServerIP, 139)
report_file = open(report_file_path, 'rb')
conn.storeFile('share_name', report_network_path, report_file)
report_file.close()
conn.close()

copyFileDuration = timeit.default_timer() - copyFileTimerStart
print('Copy File Time: ' + str(copyFileDuration) + ' ' + report_network_path)

if (args.Debug):
    print('[+] Debug - Report file copied to network share')

VSLogout()